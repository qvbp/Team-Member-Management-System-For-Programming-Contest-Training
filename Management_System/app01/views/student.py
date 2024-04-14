from datetime import date

from django.http import JsonResponse
from django.shortcuts import render, redirect, HttpResponse
from openpyxl import load_workbook

from app01.utils.pageination import Pageination
from app01 import models
from app01.utils.form import StudentModelForm
from django.utils import timezone
from datetime import timedelta


global id_x


def student_list(request):
    """ 学生管理 """

    # 生成300条数据
    # for i in range(100):
    #     models.StudentInfo.objects.create(name="91X先生", gender=1, class_name="软工2001班", student_id="20201003034",
    #                                       CodeForce_id="akccc", AtCoder_id="akccc", VJudge_id="akccc",
    #                                       NowCoder_id="akccc", LuoGu_id="akccc", LeetCode_id="akccc",
    #                                       )

    # 搜索
    # 实现搜索框
    data_dict = {}
    search_data = request.GET.get('q', "")
    if search_data:
        data_dict["name__contains"] = search_data

    # 根据搜索条件去数据库中获取
    query_set = models.StudentInfo.objects.filter(**data_dict)

    page_object = Pageination(request, query_set, page_size=20)
    context = {
        'query_set': page_object.page_query_set,
        'page_string': page_object.html(),
    }
    """
    #用python的语法获取数据
    for obj in query_set:
        obj.gender #1/2
        obj.get_gender_display() #语法：get_字段名称_display()
        obj.depart_id  # 获取数据库中存储的那个字段值
        obj.depart.title # 根据id自动去关联的表中获取那一行的数据depart对象
    """

    return render(request, 'student_list.html', context)


def student_add(request):
    """   添加学生（ModelForm版本）   """
    title = "添加学生"
    if request.method == 'GET':
        form = StudentModelForm()
        return render(request, 'change.html', {"form": form, "title": title})

    # 用户以post提交数据
    form = StudentModelForm(data=request.POST)

    if form.is_valid():
        # 如果数据合法，保存到数据库中
        # print(form.cleaned_data)
        # 默认保存的是用户输入的所有数据，如果想要在用户输入以外增加一点值
        # form.instance.字段名 = 值
        form.save()
        student = models.StudentInfo.objects.get(name=form.cleaned_data['name'])
        models.CfInfo.objects.create(name=student, gender=student.gender, CodeForce_id=student.CodeForce_id)
        models.AtCoderInfo.objects.create(name=student, gender=student.gender, AtCoder_id=student.AtCoder_id)
        models.VJudgeInfo.objects.create(name=student, gender=student.gender, VJudge_id=student.VJudge_id)
        models.NewCodeInfo.objects.create(name=student, gender=student.gender, NewCoder_id=student.NewCoder_id,
                                          NewCoder_num=student.NewCoder_num)
        models.LuoGuInfo.objects.create(name=student, gender=student.gender, LuoGu_id=student.LuoGu_id,
                                        LuoGu_num=student.LuoGu_num)
        models.StrengthInfo.objects.create(name=student, gender=student.gender)
        return redirect('/student/list/')

    # 校验失败  （在页面上显示错误信息）
    return render(request, 'change.html', {"form": form, "title": title})


def student_edit(request, nid):
    """ 编辑学生 """
    # print(f"nid: {nid}")  # 添加这一行进行调试
    title = "编辑学生信息"
    row_object = models.StudentInfo.objects.filter(id=nid).first()
    if request.method == 'GET':
        # 根据id去数据库获取要编辑的哪一行的数据(对象)
        # instance 表示你当前操作的对象
        form = StudentModelForm(instance=row_object)
        return render(request, 'change.html', {"form": form, "title": title})

    # instance 表示你当前操作的对象，同时加上‘data=’可以进行后续的数据校验
    form = StudentModelForm(data=request.POST, instance=row_object)

    if form.is_valid():
        form.save()
        student = models.StudentInfo.objects.get(name=form.cleaned_data['name'])
        models.CfInfo.objects.filter(name=student.id).update(CodeForce_id=student.CodeForce_id)
        models.AtCoderInfo.objects.filter(name=student.id).update(AtCoder_id=student.AtCoder_id)
        models.VJudgeInfo.objects.filter(name=student.id).update(VJudge_id=student.VJudge_id)
        models.NewCodeInfo.objects.filter(name=student.id).update(NewCoder_id=student.NewCoder_id,
                                          NewCoder_num=student.NewCoder_num)
        models.LuoGuInfo.objects.filter(name=student.id).update(LuoGu_id=student.LuoGu_id,
                                        LuoGu_num=student.LuoGu_num)
        return redirect('/student/list/')
    return render(request, 'change.html', {"form": form, "title": title})


def student_delete(request, nid):
    """ 删除学生 """
    models.StudentInfo.objects.filter(id=nid).delete()
    return redirect('/student/list/')


def student_multi(request):
    """ 批量上传  """
    if request.method == 'GET':
        return render(request, 'student_multi_add.html')
    # 1.获取用户上传的文件对象
    file_object = request.FILES.get("exc")
    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):  # 表示第一行不取，从第二行开始取
        text = row[0].value  # 列是从0开始计数的
        exists = models.StudentInfo.objects.filter(name=text).exists()
        if not exists:
            data = []
            for cell in row:
                if cell.value is None:
                    cell.value = "Null"
                data.append(cell.value)
            if data[1] == "男":
                data[1] = 1
            else:
                data[1] = 2  # 下面是将数据存入数据库中
            models.StudentInfo.objects.create(name=data[0], gender=data[1], class_name=data[2], student_id=data[3],
                                              CodeForce_id=data[4], AtCoder_id=data[5], VJudge_id=data[6],
                                              NewCoder_id=data[7], NewCoder_num=data[8], LuoGu_id=data[9],
                                              LuoGu_num=data[10])

            student = models.StudentInfo.objects.get(name=data[0])
            models.CfInfo.objects.create(name=student, gender=student.gender, CodeForce_id=student.CodeForce_id)
            models.AtCoderInfo.objects.create(name=student, gender=student.gender, AtCoder_id=student.AtCoder_id)
            models.VJudgeInfo.objects.create(name=student, gender=student.gender, VJudge_id=student.VJudge_id)
            models.NewCodeInfo.objects.create(name=student, gender=student.gender, NewCoder_id=student.NewCoder_id,
                                              NewCoder_num=student.NewCoder_num)
            models.LuoGuInfo.objects.create(name=student, gender=student.gender, LuoGu_id=student.LuoGu_id,
                                            LuoGu_num=student.LuoGu_num)
            models.StrengthInfo.objects.create(name=student, gender=student.gender)
        else:
            continue

    return redirect('/student/list/')


def student_enable(request, nid):
    """ 启用学生 """
    models.StudentInfo.objects.filter(id=nid).update(is_active=True)
    return redirect('/student/list/')


def student_disable(request, nid):
    """ 停用学生 """
    models.StudentInfo.objects.filter(id=nid).update(is_active=False)
    return redirect('/student/list/')


def student_index(request, nid):
    """
    学生个人主页
    姓名  性别   总做题数
    cf：rank分数   最高rank分数    总做题数
    at：rank分数   最高rank分数    总比赛场数
    nc：rank分数   rating排名     总做题数
    vj：总做题数
    lg：rating排名 总做题数
    可视化： 总做题数（柱状图）  最近做题数（近1天，7天，30天，折线）

    """

    # 将nid变成全局变量，方便后续可视化
    global id_x
    id_x = nid

    # 注意拿到的nid是strength表中的id, 拿到的name属性其实是StudentInfo表的一个object
    name = models.StrengthInfo.objects.filter(id=nid).first().name
    strength = models.StrengthInfo.objects.filter(id=nid).first()
    student = models.StudentInfo.objects.filter(name=name.name).first()
    cf = models.CfInfo.objects.filter(name=name).first()
    at = models.AtCoderInfo.objects.filter(name=name).first()
    nc = models.NewCodeInfo.objects.filter(name=name).first()
    vj = models.VJudgeInfo.objects.filter(name=name).first()
    lg = models.LuoGuInfo.objects.filter(name=name).first()

    context = {
        'strength': strength,
        'student': student,
        'cf': cf,
        'at': at,
        'nc': nc,
        'vj': vj,
        'lg': lg,
    }

    return render(request, 'personal_detail.html', context)


def student_solved_bar(request):
    """ 构造学生主页 做题数柱状图的数据 """
    # 数据库中获取数据
    name = models.StrengthInfo.objects.filter(id=id_x).first().name
    cf = models.CfInfo.objects.filter(name=name).first()
    at = models.AtCoderInfo.objects.filter(name=name).first()
    nc = models.NewCodeInfo.objects.filter(name=name).first()
    vj = models.VJudgeInfo.objects.filter(name=name).first()
    lg = models.LuoGuInfo.objects.filter(name=name).first()
    data = [cf.solved_all_time, at.Rated_Matches, nc.solved, vj.Overall_solved, lg.solved]
    x_axis = ['CodeForce', 'AtCoder', '牛客OJ', 'VJudge', '洛谷OJ']

    legend = ['做题数/比赛场数']
    series_list = [
        {
            "name": '做题数/比赛场数',
            "type": 'bar',
            "data": data,
        },
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'series_list': series_list,
            'x_axis': x_axis,
        }
    }
    return JsonResponse(result)


def student_solved_last_seven_days_bar(request):
    """ 构造学生七天做题数据图 """
    name = models.StrengthInfo.objects.filter(id=id_x).first().name
    today = timezone.now().date()
    x_axis = []
    data = []
    legend = ['做题数']
    for i in range(7):
        date_now = today - timedelta(days=i)
        date_pass = today - timedelta(days=i+1)
        # print(date_now, date_pass)
        today_solved = 0
        yesterday_solved = 0
        exists_today = models.StudentRecord.objects.filter(student=name, date=date_now).exists()
        exists_yesterday = models.StudentRecord.objects.filter(student=name, date=date_pass).exists()
        if exists_today:
            today_solved = models.StudentRecord.objects.filter(student=name, date=date_now).first().solved_today
        if exists_yesterday:
            yesterday_solved = models.StudentRecord.objects.filter(student=name, date=date_pass).first().solved_today
        solved = today_solved - yesterday_solved
        if solved < 0:
            solved = 0
        data.append(solved)
        # print(data)
        x_axis.append(date_pass)

    series_list = [
        {
            "name": '最近七天做题数',
            "type": 'bar',
            "data": data,
        },
    ]

    result = {
        "status": True,
        "data": {
            'legend': legend,
            'series_list': series_list,
            'x_axis': x_axis,
        }
    }
    return JsonResponse(result)






